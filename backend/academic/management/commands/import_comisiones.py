"""
Management command para importar comisiones desde CSV/Excel.

Uso:
    python manage.py import_comisiones ruta/al/archivo.csv
    python manage.py import_comisiones ruta/al/archivo.xlsx --dry-run

Formatos soportados:
    - CSV (separado por comas)
    - Excel (.xlsx)

Columnas esperadas:
    - Período lectivo
    - Actividad (código opcional + nombre)
    - Comisión (código)
    - Modalidad (Presencial/Remota/Híbrida) o Sede (para práctico profesional)
    - Docente (nombre completo)
    - Horario
    - RECOMENDACIÓN (opcional)
    - Sede (opcional, útil para práctico profesional y centros externos)
    - Ciclo (CPO/CPC) → ahora se requiere indicar el ciclo en la ejecución
"""
import csv
import re
from pathlib import Path
from collections import defaultdict
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from academic.models import Docente, Comision


class Command(BaseCommand):
    help = 'Importa comisiones y docentes desde un archivo CSV o Excel'

    ORIENTACIONES_INTERNAS = {
        'GENERAL', 'PENAL', 'NOTARIAL', 'AMBIENTAL', 'INT.PUBLICO', 'INT. PUBLICO',
        'INMIG./REFUG.', 'VIOLENCIA DE GENERO', 'VIOLENCIA DE GÉNERO', 'CIVIL', 'EMPRESARIAL'
    }

    def add_arguments(self, parser):
        parser.add_argument(
            'file_path',
            type=str,
            help='Ruta al archivo CSV o Excel a importar'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Simula la importación sin guardar en la base de datos'
        )
        parser.add_argument(
            '--update-existing',
            action='store_true',
            help='Actualiza comisiones existentes en lugar de omitirlas'
        )
        parser.add_argument(
            '--ciclo',
            type=str,
            choices=['CPO', 'CPC'],
            required=False,
            help='Ciclo al que pertenecen las comisiones a importar (CPO o CPC). Obligatorio cuando se ejecuta.'
        )

    def handle(self, *args, **options):
        # Reset previous run info for programmatic callers
        self.last_run_result = None
        file_path = Path(options['file_path'])
        dry_run = options['dry_run']
        update_existing = options['update_existing']
        ciclo = (options.get('ciclo') or '').strip().upper()

        if ciclo and ciclo not in {'CPO', 'CPC'}:
            raise CommandError('El ciclo debe ser CPO o CPC si se especifica.')

        self.ciclo_import = ciclo

        # Verificar que el archivo existe
        if not file_path.exists():
            raise CommandError(f'El archivo {file_path} no existe')

        # Determinar el tipo de archivo
        if file_path.suffix.lower() == '.csv':
            data = self.read_csv(file_path)
        elif file_path.suffix.lower() in ['.xlsx', '.xls']:
            data = self.read_excel(file_path)
        else:
            raise CommandError(f'Formato no soportado: {file_path.suffix}')

        # Estadísticas
        stats = {
            'docentes_creados': 0,
            'docentes_existentes': 0,
            'comisiones_creadas': 0,
            'comisiones_actualizadas': 0,
            'comisiones_omitidas': 0,
            'duplicados_exactos_omitidos': 0,
            'variaciones_detectadas': 0,
            'errores': 0,
        }

        # Detectar duplicados en el archivo
        dup_info = self.detect_duplicates(data)
        
        # Mostrar advertencias
        if dup_info['warnings']:
            self.stdout.write('\n⚠️  DUPLICADOS DETECTADOS EN EL ARCHIVO:')
            for warning in dup_info['warnings']:
                self.stdout.write(f"   {warning}")
            self.stdout.write('')

        with transaction.atomic():
            procesadas = set()  # Rastrear comisiones ya procesadas (código + docente + horario)
            
            for row in data:
                # Verificar si esta combinación ya fue procesada
                codigo_comision = (row.get('Comisión') or '').strip()
                docente_nombre = (row.get('Docente') or '').strip()
                horario = (row.get('Horario') or '').strip()
                sede = (row.get('Sede') or '').strip()
                
                # Identificador único: código + docente + horario + sede
                identificador = f"{codigo_comision}|{docente_nombre}|{horario}|{sede}"
                
                if identificador and identificador in dup_info['exactos']:
                    if identificador in procesadas:
                        # Ya procesamos esta combinación exacta
                        stats['duplicados_exactos_omitidos'] += 1
                        continue
                    procesadas.add(identificador)
                elif codigo_comision and codigo_comision in dup_info['variaciones']:
                    # Para variaciones, permitir que se procesen (múltiples horarios válidos)
                    if identificador in procesadas:
                        stats['variaciones_detectadas'] += 1
                        continue
                    procesadas.add(identificador)
                
                result = self.process_row(row, update_existing)
                if result:
                    for key, value in result.items():
                        if key in stats:
                            stats[key] += value

            # Si es dry-run, hacer rollback
            if dry_run:
                transaction.set_rollback(True)

        # Consolidar duplicados: mismo código + docente + cuatrimestre + sede
        # nos quedamos con el horario más descriptivo (string más largo)
        # para evitar fichas duplicadas con horarios parciales en el front.
        if not dry_run:
            self.cleanup_comisiones_duplicadas()

        # Guardar resultado para usos programáticos (API, tests, etc.)
        self.last_run_result = {
            'stats': stats,
            'duplicates': dup_info,
            'dry_run': dry_run,
        }

        # Mostrar resumen
        self.print_summary(stats, dry_run)

    def read_csv(self, file_path):
        """Lee un archivo CSV y retorna una lista de diccionarios."""
        self.stdout.write(f'📄 Leyendo CSV: {file_path}')
        
        # Intentar diferentes encodings
        encodings = ['utf-8-sig', 'utf-8', 'iso-8859-1', 'cp1252']
        data = None
        encoding_usado = None
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    # Leer primeras líneas para detectar headers
                    lines = f.readlines()
                    
                    # Buscar la línea que contiene "Período lectivo" (es el header)
                    header_line_idx = None
                    for i, line in enumerate(lines):
                        if 'Período lectivo' in line or 'Periodo lectivo' in line:
                            header_line_idx = i
                            break
                    
                    if header_line_idx is None:
                        # Si no encuentra el header, asumir que está en línea 0
                        header_line_idx = 0
                
                # Ahora leer correctamente
                with open(file_path, 'r', encoding=encoding) as f:
                    # Saltar líneas hasta el header
                    for _ in range(header_line_idx):
                        f.readline()
                    
                    # Detectar el delimitador
                    sample = f.read(1024)
                    f.seek(0)
                    for _ in range(header_line_idx):
                        f.readline()
                    
                    delimiter = ',' if sample.count(',') > sample.count(';') else ','
                    
                    reader = csv.DictReader(f, delimiter=delimiter)
                    data = list(reader)
                    encoding_usado = encoding
                    break
                    
            except (UnicodeDecodeError, UnicodeError):
                continue
        
        if data is None:
            raise CommandError(f'No se pudo leer el archivo con ningún encoding')
        
        self.stdout.write(f'✅ {len(data)} filas leídas (encoding: {encoding_usado})\n')
        return data

    def read_excel(self, file_path):
        """Lee un archivo Excel (xlsx/xls) y retorna una lista de diccionarios."""
        suffix = file_path.suffix.lower()
        self.stdout.write(f'📊 Leyendo Excel: {file_path}')

        if suffix == '.xls':
            try:
                import xlrd  # type: ignore[import-not-found]
            except ImportError as exc:
                raise CommandError(
                    'Para importar archivos .xls, instala xlrd:\n'
                    'pip install xlrd'
                ) from exc

            book = xlrd.open_workbook(file_path)
            sheet = book.sheet_by_index(0)
            headers = sheet.row_values(0)
            data = []
            for row_idx in range(1, sheet.nrows):
                row_values = sheet.row_values(row_idx)
                data.append(dict(zip(headers, row_values)))
            self.stdout.write(f'✅ {len(data)} filas leídas\n')
            return data

        try:
            import openpyxl  # type: ignore[import-not-found]
        except ImportError as exc:
            raise CommandError(
                'Para importar archivos .xlsx, instala openpyxl:\n'
                'pip install openpyxl'
            ) from exc

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Primera fila son los headers
        headers = [cell.value for cell in ws[1]]

        # Convertir a lista de diccionarios
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))

        self.stdout.write(f'✅ {len(data)} filas leídas\n')
        return data

    def detect_duplicates(self, data):
        """
        Detecta duplicados dentro del mismo archivo.
        
        Identificador único = código comisión + docente + horario
        Esto permite múltiples horarios para la misma comisión (VÁLIDO)
        pero detecta duplicados exactos (INVÁLIDO).
        
        Retorna diccionario con:
        - exactos: combinaciones (código|docente|horario) que aparecen múltiples veces
        - variaciones: misma comisión pero diferente docente o horario
        - warnings: lista de mensajes de advertencia
        - errores: problemas serios (ej: múltiples docentes)
        """
        # Agrupar por identificador único completo
        duplicados_por_id = defaultdict(list)
        comisiones_por_codigo = defaultdict(set)
        resultado = {'exactos': set(), 'variaciones': set(), 'warnings': [], 'errores': []}
        
        for idx, row in enumerate(data, 1):
            codigo = (row.get('Comisión') or '').strip()
            docente = (row.get('Docente') or '').strip()
            horario = (row.get('Horario') or '').strip()
            modalidad = (row.get('Modalidad') or '').strip()
            sede = (row.get('Sede') or '').strip()
            actividad = (row.get('Actividad') or '').strip()
            
            if not codigo:
                continue
            
            # ID único para esta instancia de comisión
            id_unico = f"{codigo}|{docente}|{horario}|{sede}"
            
            # Registrar para análisis de duplicados exactos
            duplicados_por_id[id_unico].append({
                'fila': idx,
                'docente': docente,
                'horario': horario,
                'modalidad': modalidad,
                'actividad': actividad,
            })
            
            # Registrar docentes para esta comisión (detectar múltiples docentes)
            comisiones_por_codigo[codigo].add(docente)
        
        # ANÁLISIS 1: Detectar duplicados exactos (mismo ID unico aparece múltiples veces)
        for id_unico, instancias in duplicados_por_id.items():
            if len(instancias) > 1:
                resultado['exactos'].add(id_unico)
                codigo = id_unico.split('|')[0]
                filas = [str(inst['fila']) for inst in instancias]
                resultado['warnings'].append(
                    f"Comisión {codigo} (docente: {instancias[0]['docente']}, horario: {instancias[0]['horario']}, sede: {sede or 'N/D'}) "
                    f"aparece {len(instancias)} veces (idénticas, filas: {', '.join(filas)})"
                )
        
        # ANÁLISIS 2: Detectar si misma comisión tiene múltiples docentes (ERROR)
        for codigo, docentes in comisiones_por_codigo.items():
            if len(docentes) > 1:
                resultado['errores'].append(
                    f"Comisión {codigo} asignada a múltiples docentes: {', '.join(sorted(docentes))}"
                )
                # Marcar como variación para mostrar advertencia
                resultado['variaciones'].add(codigo)
        
        # ANÁLISIS 3: Detectar variaciones (mismo código, diferente horario/docente, pero válido)
        for codigo, docentes in comisiones_por_codigo.items():
            if len(docentes) == 1:
                # Un solo docente, revisar si hay múltiples horarios
                docente_unico = list(docentes)[0]
                horarios = set()
                for id_unico in duplicados_por_id.keys():
                    partes = id_unico.split('|')
                    if partes[0] == codigo and partes[1] == docente_unico:
                        horarios.add(partes[2])
                
                if len(horarios) > 1:
                    resultado['variaciones'].add(codigo)
                    resultado['warnings'].append(
                        f"Comisión {codigo} ({docente_unico}) tiene {len(horarios)} horarios diferentes: "
                        f"{', '.join(sorted(horarios))}"
                    )
        
        return resultado

    def process_row(self, row, update_existing):
        """
        Procesa una fila del CSV/Excel y crea/actualiza los modelos.
        
        Retorna un diccionario con estadísticas de la operación.
        """
        stats = {
            'docentes_creados': 0,
            'docentes_existentes': 0,
            'comisiones_creadas': 0,
            'comisiones_actualizadas': 0,
            'comisiones_omitidas': 0,
        }

        # 1. PARSEAR DOCENTE
        docente_nombre_completo = (row.get('Docente') or '').strip()
        if not docente_nombre_completo:
            return stats  # Omitir filas sin docente

        # Separar nombre y apellido
        # Asumimos formato: "APELLIDO NOMBRE" o "APELLIDO NOMBRE1 NOMBRE2"
        partes = docente_nombre_completo.split()
        if len(partes) >= 2:
            apellido = partes[0]
            nombre = ' '.join(partes[1:])
        else:
            apellido = docente_nombre_completo
            nombre = ''

        # Buscar o crear docente
        docente, created = Docente.objects.get_or_create(
            nombre_completo__iexact=docente_nombre_completo,
            defaults={
                'nombre': nombre.title(),
                'apellido': apellido.title(),
                'nombre_completo': docente_nombre_completo.title(),
            }
        )

        if created:
            stats['docentes_creados'] += 1
            self.stdout.write(f'  👤 Docente creado: {docente.nombre_completo}')
        else:
            stats['docentes_existentes'] += 1

        # 2. PARSEAR COMISIÓN
        actividad = (row.get('Actividad') or '').strip()
        codigo_comision = (row.get('Comisión') or '').strip()
        
        if not actividad or not codigo_comision:
            return stats  # Omitir filas incompletas

        # Extraer código de actividad (si existe)
        # Formato: "205 (PRI) - DERECHO ROMANO" o "DERECHO ROMANO"
        match = re.match(r'^(\S+)\s+\([^)]+\)\s*-\s*(.+)$', actividad)
        if match:
            codigo_actividad = match.group(1)
            nombre_actividad = match.group(2).strip()
        else:
            codigo_actividad = ''
            nombre_actividad = actividad

        # Modalidad y sede (para práctico profesional)
        modalidad = (row.get('Modalidad') or '').strip()
        sede = (row.get('Sede') or '').strip()
        es_centro_externo = self.is_centro_externo(sede, row)
        ciclo = self.ciclo_import
        
        # Horario
        horario = (row.get('Horario') or '').strip()
        
        # Recomendación (raw, sin procesar)
        recomendacion_raw = (row.get('RECOMENDACIÓN') or '').strip()
        
        # Período lectivo (para extraer cuatrimestre)
        periodo = (row.get('Período lectivo') or '').strip()
        cuatrimestre = self.extract_cuatrimestre(periodo)

        # El identificador único CORRECTO es: código + docente + horario + cuatrimestre
        # Esto permite múltiples horarios para la misma comisión
        
        # Consolidar por código+docente+cuatrimestre+sede: mantener horario más descriptivo
        existentes = list(Comision.objects.filter(
            codigo=codigo_comision,
            docente=docente,
            cuatrimestre=cuatrimestre,
            sede=sede,
        ))

        # Si no se quiere actualizar y ya hay uno, omitir
        if existentes and not update_existing:
            stats['comisiones_omitidas'] += 1
            return stats

        # Elegir target (máxima longitud de horario entre existentes y el nuevo)
        if existentes:
            candidato = max(existentes + [Comision(horario=horario or '')], key=lambda c: len(c.horario or ''))
            target = candidato if candidato.id_comision else existentes[0]
            target.horario = candidato.horario or horario
            created = False
        else:
            target = Comision(
                codigo=codigo_comision,
                docente=docente,
                cuatrimestre=cuatrimestre,
                sede=sede,
                horario=horario,
            )
            created = True

        # Actualizar campos comunes
        target.codigo_actividad = codigo_actividad
        target.nombre = nombre_actividad[:200]
        target.modalidad = modalidad if modalidad in ['Presencial', 'Remota', 'Híbrida'] else None
        target.recomendacion_raw = recomendacion_raw
        target.sede = sede
        target.es_centro_externo = es_centro_externo
        target.ciclo = ciclo
        target.activa = True
        target.save()

        # Eliminar duplicados restantes si los hay
        if existentes:
            for item in existentes:
                if item.id_comision != target.id_comision:
                    item.delete()

        if created:
            stats['comisiones_creadas'] += 1
            self.stdout.write(
                f'  ✅ Comisión creada: {codigo_comision} - {nombre_actividad[:50]}...'
            )
        else:
            stats['comisiones_actualizadas'] += 1
            self.stdout.write(f'  📝 Comisión actualizada: {codigo_comision}')

        return stats

    def extract_cuatrimestre(self, periodo_text):
        """
        Extrae el cuatrimestre del texto del período.
        
        Ejemplos:
            "PRIMER CUATRIMESTRE ABOGACÍA 2025" -> "1C2025"
            "SEGUNDO BIMESTRE ABOGACÍA 2025" -> "2B2025"
        """
        # Buscar año
        year_match = re.search(r'20\d{2}', periodo_text)
        year = year_match.group(0) if year_match else ''
        
        # Buscar cuatrimestre/bimestre
        if 'PRIMER' in periodo_text.upper():
            periodo = '1'
        elif 'SEGUNDO' in periodo_text.upper():
            periodo = '2'
        else:
            periodo = '1'
        
        if 'CUATRIMESTRE' in periodo_text.upper():
            tipo = 'C'
        elif 'BIMESTRE' in periodo_text.upper():
            tipo = 'B'
        else:
            tipo = 'C'
        
        return f"{periodo}{tipo}{year}" if year else ''

    def is_centro_externo(self, sede_value, row):
        """Determina si la comisión corresponde a un centro externo.

        Reglas:
        - Si existe una columna 'Centro externo'/'Centros externos' con valor afirmativo → True
        - Si la sede no está en la lista de orientaciones internas conocidas → probablemente externo
        """
        afirmativo = {'si', 'sí', 'yes', 'true', '1', 'externo', 'externa'}
        centro_col = (row.get('Centro externo') or row.get('Centros externos') or '').strip().lower()
        if centro_col in afirmativo:
            return True

        sede_norm = (sede_value or '').strip().upper()
        if not sede_norm:
            return False

        if sede_norm in self.ORIENTACIONES_INTERNAS:
            return False

        # Si la sede es un nombre institucional (colegio, defensoría, universidad, etc.) lo marcamos como externo
        palabras_externas = ['COLEGIO', 'DEFENSORIA', 'UNIVERSIDAD', 'HOSPITAL', 'TRIBUNAL', 'CENTRO']
        if any(p in sede_norm for p in palabras_externas):
            return True

        # Por defecto, si no es orientación conocida, lo tratamos como externo
        return True

    # -------------------------------------------------------------
    # Limpieza de duplicados (mismo código+docente+cuatrimestre+sede)
    # -------------------------------------------------------------
    def cleanup_comisiones_duplicadas(self):
        from academic.models import Comision  # Import local para evitar ciclos

        grupos = {}
        for com in Comision.objects.all():
            key = (com.codigo, com.docente, com.cuatrimestre or '', com.sede or '')
            grupos.setdefault(key, []).append(com)

        eliminadas = 0
        for _key, items in grupos.items():
            if len(items) <= 1:
                continue

            items_sorted = sorted(items, key=lambda c: len(c.horario or ''), reverse=True)
            keep = items_sorted[0]
            to_delete = items_sorted[1:]

            ids = [c.id_comision for c in to_delete]
            if ids:
                Comision.objects.filter(id_comision__in=ids).delete()
                eliminadas += len(ids)

        if eliminadas:
            self.stdout.write(f"🧹 Comisiones duplicadas consolidadas: {eliminadas}")

    def print_summary(self, stats, dry_run):
        """Imprime un resumen de la importación."""
        self.stdout.write('\n' + '='*60)
        self.stdout.write('📊 RESUMEN DE IMPORTACIÓN')
        self.stdout.write('='*60 + '\n')
        
        if dry_run:
            self.stdout.write('⚠️  Modo DRY-RUN (no se guardó nada)\n')
        
        self.stdout.write("👤 Docentes:")
        self.stdout.write(f"   • Creados: {stats['docentes_creados']}")
        self.stdout.write(f"   • Ya existentes: {stats['docentes_existentes']}")
        self.stdout.write('')
        
        self.stdout.write("📚 Comisiones:")
        self.stdout.write(f"   • Creadas: {stats['comisiones_creadas']}")
        self.stdout.write(f"   • Actualizadas: {stats['comisiones_actualizadas']}")
        self.stdout.write(f"   • Omitidas: {stats['comisiones_omitidas']}")
        self.stdout.write('')
        
        # Mostrar estadísticas de duplicados
        if stats['duplicados_exactos_omitidos'] > 0 or stats['variaciones_detectadas'] > 0:
            self.stdout.write("🔄 Duplicados procesados:")
            if stats['duplicados_exactos_omitidos'] > 0:
                self.stdout.write(f"   • Exactos omitidos: {stats['duplicados_exactos_omitidos']}")
            if stats['variaciones_detectadas'] > 0:
                self.stdout.write(f"   • Variaciones procesadas: {stats['variaciones_detectadas']}")
            self.stdout.write('')
        
        if stats['errores'] > 0:
            self.stdout.write(f"❌ Errores: {stats['errores']}")
        else:
            self.stdout.write("✅ Sin errores")
        
        self.stdout.write('\n' + '='*60)
